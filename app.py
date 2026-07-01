import csv
import hashlib
import io
import json
import logging
import os
import re
import secrets
import sqlite3
import time

logging.basicConfig(level=logging.INFO)

import anthropic
from flask import Flask, Response, jsonify, redirect, render_template, request, session, url_for
from functools import wraps

from analyzer import analyze_niche
from meli_api import (
    fetch_orders_total, get_categories, get_my_store_items, get_my_user_id,
    get_seller_info, get_subcategories, resolve_seller_alias, sample_subcategory,
    search_meli, snapshot_seller,
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 120 * 1024 * 1024  # 120 MB
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "melichnicho.db")


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS searches (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                query       TEXT,
                category_id TEXT,
                filters     TEXT,
                results_count INTEGER,
                niche_stats TEXT,
                timestamp   DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS search_results (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                search_id        INTEGER,
                item_id          TEXT,
                title            TEXT,
                price            REAL,
                sold_quantity    INTEGER,
                seller_id        TEXT,
                seller_name      TEXT,
                opportunity_score REAL,
                free_shipping    INTEGER,
                permalink        TEXT,
                thumbnail        TEXT,
                FOREIGN KEY (search_id) REFERENCES searches(id)
            );
        """)


def init_users_table():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        admin_user = os.environ.get("ADMIN_USER")
        admin_pass = os.environ.get("ADMIN_PASSWORD")
        logging.info(f"ADMIN_USER env var present: {bool(admin_user)}, ADMIN_PASSWORD env var present: {bool(admin_pass)}")
        if admin_user and admin_pass:
            pw_hash = hashlib.sha256(admin_pass.encode()).hexdigest()
            logging.info(f"Upserting admin user: '{admin_user}'")
            conn.execute(
                "INSERT INTO users (username, password_hash) VALUES (?, ?)"
                " ON CONFLICT(username) DO UPDATE SET password_hash = excluded.password_hash",
                (admin_user, pw_hash),
            )
            count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            logging.info(f"Users in DB: {count}")


def init_competitors_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS competitors (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                seller_id   TEXT UNIQUE NOT NULL,
                nickname    TEXT,
                permalink   TEXT,
                added_at    DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS competitor_snapshots (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                competitor_id INTEGER NOT NULL,
                taken_at      DATETIME DEFAULT CURRENT_TIMESTAMP,
                item_count    INTEGER,
                seller_json   TEXT,
                items_json    TEXT,
                FOREIGN KEY (competitor_id) REFERENCES competitors(id)
            );
            -- Estado de snapshots en curso. Vive en SQLite (no en memoria) porque
            -- Gunicorn corre 2 workers y el polling puede caer en cualquiera.
            CREATE TABLE IF NOT EXISTS competitor_jobs (
                competitor_id INTEGER PRIMARY KEY,
                status        TEXT,
                pct           INTEGER,
                msg           TEXT,
                error         TEXT,
                updated_at    DATETIME DEFAULT CURRENT_TIMESTAMP
            );
        """)


init_db()
init_users_table()
init_competitors_db()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "No autorizado"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        pw_hash = hashlib.sha256(password.encode()).hexdigest()
        with sqlite3.connect(DB_PATH) as conn:
            row = conn.execute(
                "SELECT id FROM users WHERE username = ? AND password_hash = ?",
                (username, pw_hash),
            ).fetchone()
        logging.info(f"Login attempt: user='{username}', hash='{pw_hash[:12]}...'")
        if row:
            session["user"] = username
            logging.info(f"Login success: {username}")
            return redirect(url_for("index"))
        logging.info(f"Login failed: no matching user/password")
        error = "Usuario o contraseña incorrectos"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/nubi-results")
@login_required
def nubi_results():
    return render_template("nubi_results.html")


@app.route("/api/categories")
@login_required
def categories():
    return jsonify(get_categories())


@app.route("/api/search", methods=["POST"])
@login_required
def search():
    data = request.get_json() or {}
    query = (data.get("query") or "").strip()
    category_id = (data.get("category_id") or "").strip()

    if not query and not category_id:
        return jsonify({"error": "Ingresá una keyword o seleccioná una categoría."}), 400

    filters = {
        "min_price": data.get("min_price"),
        "max_price": data.get("max_price"),
        "min_sold": data.get("min_sold", 0),
        "max_sellers": data.get("max_sellers"),
        "free_shipping": data.get("free_shipping", False),
    }

    raw = search_meli(query=query, category_id=category_id)
    if "error" in raw:
        return jsonify(raw), 502

    items, niche_stats, seller_ranking = analyze_niche(raw["items"], filters)

    max_sellers = filters.get("max_sellers")
    if max_sellers and niche_stats.get("unique_sellers", 0) > int(max_sellers):
        niche_stats["competition_warning"] = (
            f"Nicho con {niche_stats['unique_sellers']} vendedores "
            f"(tu límite: {int(max_sellers)})"
        )

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO searches (query, category_id, filters, results_count, niche_stats)"
            " VALUES (?,?,?,?,?)",
            (query or category_id, category_id, json.dumps(filters),
             len(items), json.dumps(niche_stats)),
        )
        search_id = c.lastrowid
        c.executemany(
            "INSERT INTO search_results"
            " (search_id, item_id, title, price, sold_quantity, seller_id, seller_name,"
            "  opportunity_score, free_shipping, permalink, thumbnail)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            [
                (search_id, i["id"], i["title"], i["price"], i.get("sold_quantity", 0),
                 str(i.get("seller_id", "")), i.get("seller_name", ""),
                 i.get("opportunity_score", 0), 1 if i.get("free_shipping") else 0,
                 i.get("permalink", ""), i.get("thumbnail", ""))
                for i in items
            ],
        )

    return jsonify({
        "search_id": search_id,
        "items": items,
        "niche_stats": niche_stats,
        "seller_ranking": seller_ranking,
    })


@app.route("/api/history")
@login_required
def history():
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, query, category_id, results_count, timestamp"
            " FROM searches ORDER BY timestamp DESC LIMIT 20"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/export/<int:search_id>")
@login_required
def export_csv(search_id):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM search_results WHERE search_id = ?"
            " ORDER BY opportunity_score DESC",
            (search_id,),
        ).fetchall()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Título", "Precio (ARS)", "Vendidos", "Vendedor",
                "Score Oportunidad", "Envío Gratis", "Link"])
    for r in rows:
        w.writerow([
            r["title"], r["price"], r["sold_quantity"], r["seller_name"],
            r["opportunity_score"], "Sí" if r["free_shipping"] else "No", r["permalink"],
        ])

    return Response(
        "﻿" + buf.getvalue(),  # BOM for Excel
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=melichnicho_{search_id}.csv"},
    )


@app.route("/api/discover", methods=["POST"])
@login_required
def discover():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    category_id = (data.get("category_id") or "").strip()
    if not category_id:
        return jsonify({"error": "Seleccioná una categoría."}), 400

    subcats = get_subcategories(category_id)
    if not subcats:
        return jsonify({"error": "No se encontraron subcategorías para analizar."}), 404

    subcats_with_data = []
    for sc in subcats[:8]:
        metrics = sample_subcategory(sc["id"])
        if metrics:
            subcats_with_data.append({
                "id": sc["id"],
                "name": sc["name"],
                "total_items_in_category": sc.get("total_items_in_this_category", 0),
                **metrics,
            })
        time.sleep(0.3)

    if not subcats_with_data:
        return jsonify({"error": "No se pudieron obtener datos de las subcategorías."}), 502

    datos_str = json.dumps(subcats_with_data, ensure_ascii=False, indent=2)
    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso datos de subcategorías dentro de una categoría.\n"
        "Analizá y encontrá las MEJORES OPORTUNIDADES de nicho, priorizando:\n"
        "1. Pocos vendedores únicos (baja competencia)\n"
        "2. Buena demanda (ventas promedio decentes)\n"
        "3. Precios que permitan margen\n\n"
        "Devolvé un informe en español con:\n"
        "- Top 3-5 nichos recomendados, ordenados de mejor a peor oportunidad\n"
        "- Para cada uno: nombre de la subcategoría, por qué es buena oportunidad, "
        "nivel de competencia, rango de precios sugerido, y un veredicto "
        "(🟢 Entrar / 🟡 Evaluar / 🔴 Evitar)\n"
        "- Un resumen final con tu recomendación principal\n\n"
        f"Datos de subcategorías:\n{datos_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = response.content[0].text
        return jsonify({
            "analysis": analysis_text,
            "subcategories_analyzed": len(subcats_with_data),
            "data": subcats_with_data,
        })
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/analyze", methods=["POST"])
@login_required
def analyze_ai():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    niche_data = data.get("niche_data", {})
    datos_str = json.dumps(niche_data, ensure_ascii=False, indent=2)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Analizá este nicho y devolvé un informe breve en español con:\n"
        "1. Resumen del mercado\n"
        "2. Nivel de competencia (bajo/medio/alto)\n"
        "3. Oportunidad de entrada\n"
        "4. Precio recomendado para entrar\n"
        "5. Veredicto final: una de estas tres opciones exactas → "
        "🟢 Recomendado / 🟡 Evaluar / 🔴 Evitar\n\n"
        f"Datos del nicho: {datos_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = response.content[0].text

        tail = analysis_text[-600:].lower()
        if "🟢" in analysis_text[-600:] or "recomendado" in tail:
            verdict = "green"
        elif "🔴" in analysis_text[-600:] or "evitar" in tail:
            verdict = "red"
        else:
            verdict = "yellow"

        return jsonify({"analysis": analysis_text, "verdict": verdict})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/rt-upload", methods=["POST"])
@login_required
def rt_upload():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl no instalado en el servidor."}), 500

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el archivo: {e}"}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "El archivo está vacío."}), 400

    # Detectar fila de encabezado buscando columnas clave
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        if any("unidades" in c for c in row_lower) or any("publicacion" in c for c in row_lower):
            header_row = row_lower
            header_idx = i
            break

    if header_row is None:
        return jsonify({"error": "No se encontraron las columnas esperadas (Unidades vendidas, Publicaciones)."}), 400

    def col_idx(keywords):
        for i, h in enumerate(header_row):
            if any(k in h for k in keywords):
                return i
        return None

    idx_title   = col_idx(["publicacion", "título", "titulo", "producto", "nombre"])
    idx_seller  = col_idx(["vendedor", "seller"])
    idx_price   = col_idx(["precio"])
    idx_units   = col_idx(["unidades"])
    idx_revenue = col_idx(["facturacion", "facturación", "revenue"])

    items = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        def val(i):
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v

        title = val(idx_title)
        if not title:
            continue

        # Limpiar número: "1.500" → 1500, "+1.500" → 1500
        def to_num(v):
            if v is None:
                return 0
            s = re.sub(r"[^\d]", "", str(v))
            return int(s) if s else 0

        def to_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).replace("$", "").replace("+", "").strip()
            s = s.replace(".", "").replace(",", ".")
            try:
                return float(s)
            except:
                return 0.0

        items.append({
            "title":    str(title).strip(),
            "seller":   str(val(idx_seller)).strip() if val(idx_seller) else "",
            "price":    to_float(val(idx_price)),
            "units":    to_num(val(idx_units)),
            "revenue":  to_float(val(idx_revenue)),
        })

    if not items:
        return jsonify({"error": "No se encontraron filas de datos en el archivo."}), 400

    items.sort(key=lambda x: x["units"], reverse=True)
    return jsonify({"items": items, "total": len(items)})


@app.route("/api/rt-analyze", methods=["POST"])
@login_required
def rt_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    items = data.get("items", [])
    if not items:
        return jsonify({"error": "No hay datos para analizar."}), 400

    items_str = json.dumps(items[:30], ensure_ascii=False, indent=2)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso el ranking de publicaciones de una categoría exportado desde Real Trends.\n"
        "Cada item tiene: título, vendedor, precio promedio, unidades vendidas, facturación.\n\n"
        "Analizá estos datos y detectá OPORTUNIDADES DE NEGOCIO para un vendedor que quiere entrar o crecer en esta categoría.\n\n"
        "Tu análisis debe incluir:\n"
        "1. **Resumen del mercado** — tamaño, concentración, quiénes dominan\n"
        "2. **Segmentos de precio** — identificá rangos de precio y cuáles tienen más demanda\n"
        "3. **Oportunidades detectadas** — productos o nichos dentro de la categoría con:\n"
        "   - Alta demanda (unidades vendidas) pero pocos competidores dominantes\n"
        "   - Segmentos de precio sin jugadores fuertes\n"
        "   - Variantes de producto sub-representadas en el top\n"
        "4. **Estrategia recomendada** — precio de entrada, tipo de producto, diferenciación\n"
        "5. **Veredicto** — 🟢 Oportunidad clara / 🟡 Evaluar / 🔴 Mercado saturado\n\n"
        f"Datos del ranking:\n{items_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/nubi-upload", methods=["POST"])
@login_required
def nubi_upload():
    import statistics
    from collections import defaultdict

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    try:
        stream = io.StringIO(f.stream.read().decode("utf-8", errors="replace"))
        reader = csv.DictReader(stream)
        rows = list(reader)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el CSV: {e}"}), 400

    if not rows:
        return jsonify({"error": "El archivo está vacío."}), 400

    def fnum(v):
        try: return float(v or 0)
        except: return 0.0

    def fint(v):
        try: return int(float(v or 0))
        except: return 0

    # ─── Aggregación por subcategoría ─────────────────────
    subcats = defaultdict(lambda: {
        "units": 0, "revenue": 0, "listings": 0,
        "prices": [], "full": 0, "fship": 0,
        "sellers": defaultdict(int),
        "products": defaultdict(lambda: {"units": 0, "price": 0, "seller": ""}),
    })

    price_buckets = {"<10k": [0,0], "10k-25k": [0,0], "25k-50k": [0,0],
                     "50k-100k": [0,0], "100k-200k": [0,0], "+200k": [0,0]}

    total_units = 0
    total_revenue = 0.0
    unique_sellers_global = set()

    for r in rows:
        cat = (r.get("Categoria_Nivel_4") or "").strip()
        if not cat or cat == "-":
            cat = (r.get("Categoria_Nivel_3") or "").strip() or "Otros"

        u    = fint(r.get("Unidades_Vendidas"))
        rev  = fnum(r.get("Monto_Vendido_Moneda_Local"))
        price = fnum(r.get("PrecioMonedaLocal"))
        seller = r.get("Nickname_Vendedor", "")
        title  = (r.get("Titulo_Publicacion") or "")[:80].strip()

        total_units   += u
        total_revenue += rev
        unique_sellers_global.add(seller)

        d = subcats[cat]
        d["units"]    += u
        d["revenue"]  += rev
        d["listings"] += 1
        if price > 0: d["prices"].append(price)
        if r.get("OfreceFull") == "Si":           d["full"]  += 1
        if r.get("Ofrece_Envio_Gratis") == "true": d["fship"] += 1
        d["sellers"][seller] += u
        p = d["products"][title]
        p["units"] += u
        if price > p["price"]: p["price"] = price
        p["seller"] = seller

        # price buckets
        if price < 10_000:      price_buckets["<10k"][0] += u;    price_buckets["<10k"][1] += 1
        elif price < 25_000:    price_buckets["10k-25k"][0] += u; price_buckets["10k-25k"][1] += 1
        elif price < 50_000:    price_buckets["25k-50k"][0] += u; price_buckets["25k-50k"][1] += 1
        elif price < 100_000:   price_buckets["50k-100k"][0] += u; price_buckets["50k-100k"][1] += 1
        elif price < 200_000:   price_buckets["100k-200k"][0] += u; price_buckets["100k-200k"][1] += 1
        else:                   price_buckets["+200k"][0] += u;    price_buckets["+200k"][1] += 1

    # ─── Serializar resultado ──────────────────────────────
    subcat_list = []
    for name, d in subcats.items():
        n = d["listings"]
        u = d["units"]
        top_sellers = sorted(d["sellers"].items(), key=lambda x: -x[1])[:3]
        top3_units  = sum(x[1] for x in top_sellers)
        top_products = sorted(d["products"].items(), key=lambda x: -x[1]["units"])[:5]

        subcat_list.append({
            "name":            name,
            "listings":        n,
            "unique_sellers":  len(d["sellers"]),
            "total_units":     u,
            "total_revenue":   round(d["revenue"]),
            "avg_price":       round(statistics.mean(d["prices"])) if d["prices"] else 0,
            "median_price":    round(statistics.median(d["prices"])) if d["prices"] else 0,
            "pct_full":        round(d["full"] / n * 100) if n else 0,
            "pct_free_ship":   round(d["fship"] / n * 100) if n else 0,
            "top3_concentration": round(top3_units / u * 100) if u else 0,
            "top_sellers":     [{"seller": s, "units": v} for s, v in top_sellers],
            "top_products":    [{"title": t, "units": p["units"], "price": round(p["price"]), "seller": p["seller"]} for t, p in top_products],
        })

    subcat_list.sort(key=lambda x: -x["total_units"])

    # top 15 productos globales
    all_products = defaultdict(lambda: {"units": 0, "price": 0, "seller": "", "cat": ""})
    for r in rows:
        title = (r.get("Titulo_Publicacion") or "")[:80].strip()
        u = fint(r.get("Unidades_Vendidas"))
        price = fnum(r.get("PrecioMonedaLocal"))
        all_products[title]["units"] += u
        if price > all_products[title]["price"]: all_products[title]["price"] = price
        all_products[title]["seller"] = r.get("Nickname_Vendedor", "")
        cat = (r.get("Categoria_Nivel_4") or r.get("Categoria_Nivel_3") or "").strip()
        all_products[title]["cat"] = cat

    top_global = sorted(all_products.items(), key=lambda x: -x[1]["units"])[:15]

    meta = {
        "category_name": (rows[0].get("Categoria_Nivel_2") or rows[0].get("Categoria_Nivel_1") or "").strip(),
        "period": (rows[0].get("Mes") or "")[:7],
        "total_listings": len(rows),
        "total_units": total_units,
        "total_revenue_ars": round(total_revenue),
        "unique_sellers": len(unique_sellers_global),
    }

    return jsonify({
        "meta": meta,
        "subcategories": subcat_list,
        "top_products": [{"title": t, **p} for t, p in top_global],
        "price_segments": {k: {"units": v[0], "listings": v[1]} for k, v in price_buckets.items()},
    })


@app.route("/api/nubi-analyze", methods=["POST"])
@login_required
def nubi_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    agg  = data.get("data", {})
    if not agg:
        return jsonify({"error": "Sin datos para analizar."}), 400

    meta     = agg.get("meta", {})
    subcats  = agg.get("subcategories", [])[:20]
    top_prods = agg.get("top_products", [])
    segments = agg.get("price_segments", {})

    summary = json.dumps({
        "meta": meta,
        "subcategories": subcats,
        "top_15_productos": top_prods,
        "segmentos_de_precio": segments,
    }, ensure_ascii=False, indent=1)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso datos agregados de Nubimetrics para una categoría completa.\n\n"
        "IMPORTANTE — formato de números:\n"
        "- Expresá todos los montos en formato abreviado: M para millones, B para miles de millones.\n"
        "  Ejemplos correctos: $5.9B, $520M, $14.3M, $103k. NUNCA escribas números largos como $5.878.334.409.\n"
        "- Las unidades vendidas también abrevialas si superan mil: 14.4k, 120k.\n\n"
        "Los datos incluyen por subcategoría:\n"
        "- listings, unique_sellers, total_units, total_revenue, avg_price, median_price\n"
        "- pct_full (% publicaciones con MeLi Full), pct_free_ship, top3_concentration (% de unidades de los 3 top vendedores)\n"
        "- top_products y top_sellers\n\n"
        "Analizá estos datos y producí un informe de oportunidades de negocio con:\n\n"
        "## 1. Resumen del mercado\n"
        "Tamaño total, categorías más grandes, concentración general.\n\n"
        "## 2. Top 5 Oportunidades de Nicho\n"
        "Para cada oportunidad usá este formato exacto:\n"
        "### [Nombre de la subcategoría]\n"
        "- Por qué es oportunidad\n"
        "- Competencia: vendedores, concentración top 3\n"
        "- Precio recomendado para entrar\n"
        "- Estrategia: Full / envío gratis\n"
        "- Veredicto: 🟢 Entrar / 🟡 Evaluar / 🔴 Evitar\n\n"
        "## 3. Segmentos de precio con más demanda\n"
        "Qué rango de precio concentra más unidades y por qué.\n\n"
        "## 4. Advertencias\n"
        "Subcategorías saturadas o con barreras de entrada altas.\n\n"
        "## 5. Recomendación final\n"
        "La mejor oportunidad concreta para un vendedor nuevo o en crecimiento.\n\n"
        f"Datos:\n{summary}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/nubi-export", methods=["POST"])
@login_required
def nubi_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl no instalado."}), 500

    body = request.get_json() or {}
    data     = body.get("data", {})
    analysis = body.get("analysis", "")

    if not data:
        return jsonify({"error": "Sin datos para exportar."}), 400

    meta      = data.get("meta", {})
    subcats   = data.get("subcategories", [])
    top_prods = data.get("top_products", [])
    segments  = data.get("price_segments", {})

    wb = openpyxl.Workbook()

    # ─── Estilos ──────────────────────────────────────────
    hdr_font    = Font(bold=True, color="000000")
    hdr_fill    = PatternFill("solid", fgColor="FFE600")
    title_font  = Font(bold=True, size=13, color="FFE600")
    label_font  = Font(bold=True, color="E0E0E0")
    dark_fill   = PatternFill("solid", fgColor="16213E")
    green_fill  = PatternFill("solid", fgColor="1B5E20")
    orange_fill = PatternFill("solid", fgColor="E65100")
    red_fill    = PatternFill("solid", fgColor="B71C1C")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Border(
        left=Side(style="thin", color="0F3460"),
        right=Side(style="thin", color="0F3460"),
        top=Side(style="thin", color="0F3460"),
        bottom=Side(style="thin", color="0F3460"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = center
            cell.border = thin

    def style_data_cell(cell, fill=None):
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = thin
        if fill:
            cell.fill = fill

    # ─── Hoja 1: Resumen ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    ws1["A1"] = f"📈 Nubimetrics — {meta.get('category_name', '')} | {meta.get('period', '')}"
    ws1["A1"].font = title_font

    meta_rows = [
        ("Total de listings",    f"{meta.get('total_listings', 0):,}"),
        ("Unidades vendidas",    f"{meta.get('total_units', 0):,}"),
        ("Facturación total ARS",f"${meta.get('total_revenue_ars', 0):,.0f}"),
        ("Vendedores únicos",    f"{meta.get('unique_sellers', 0):,}"),
    ]
    for i, (label, value) in enumerate(meta_rows, start=3):
        ws1.cell(row=i, column=1, value=label).font = label_font
        ws1.cell(row=i, column=2, value=value)

    ws1["A8"] = "Segmentos de precio"
    ws1["A8"].font = Font(bold=True, color="FFE600")
    ws1["A9"]  = "Rango";        ws1["B9"]  = "Unidades";  ws1["C9"] = "Listings"
    ws1.column_dimensions["C"].width = 15
    style_header_row(ws1, 9, 3)
    for i, (k, v) in enumerate(segments.items(), start=10):
        ws1.cell(row=i, column=1, value=k)
        ws1.cell(row=i, column=2, value=v.get("units", 0))
        ws1.cell(row=i, column=3, value=v.get("listings", 0))
        for c in range(1, 4):
            style_data_cell(ws1.cell(row=i, column=c))

    # ─── Hoja 2: Subcategorías ────────────────────────────
    ws2 = wb.create_sheet("Subcategorías")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Subcategoría","Listings","Vendedores únicos","Unidades vendidas",
                 "Revenue ARS","Precio promedio","Precio mediano",
                 "% Full","% Envío gratis","Concentración Top3 %"]
    col_widths2 = [28,10,18,18,18,16,16,10,14,20]
    for i, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.cell(row=1, column=i, value=h)
    style_header_row(ws2, 1, len(headers2))

    for r, s in enumerate(subcats, start=2):
        conc = s.get("top3_concentration", 0)
        conc_fill = green_fill if conc <= 30 else (orange_fill if conc <= 50 else red_fill)
        values = [
            s.get("name",""), s.get("listings",0), s.get("unique_sellers",0),
            s.get("total_units",0), s.get("total_revenue",0),
            s.get("avg_price",0), s.get("median_price",0),
            s.get("pct_full",0), s.get("pct_free_ship",0), conc,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            style_data_cell(cell, fill=conc_fill if c == 10 else None)
            if c == 10:
                cell.font = Font(bold=True, color="FFFFFF")

    # ─── Hoja 3: Top Productos ────────────────────────────
    ws3 = wb.create_sheet("Top Productos")
    ws3.sheet_view.showGridLines = False
    headers3 = ["#","Título","Vendedor","Subcategoría","Unidades vendidas","Precio ARS"]
    col_widths3 = [5, 50, 25, 25, 18, 14]
    for i, (h, w) in enumerate(zip(headers3, col_widths3), start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, len(headers3))

    for r, p in enumerate(top_prods, start=2):
        values = [r-1, p.get("title",""), p.get("seller",""), p.get("cat",""),
                  p.get("units",0), p.get("price",0)]
        for c, v in enumerate(values, start=1):
            style_data_cell(ws3.cell(row=r, column=c, value=v))

    # ─── Hoja 4: Análisis IA ──────────────────────────────
    if analysis:
        ws4 = wb.create_sheet("Análisis IA")
        ws4.sheet_view.showGridLines = False
        ws4.column_dimensions["A"].width = 120
        # Strip markdown and write as plain text blocks
        lines = analysis.replace("**", "").replace("##", "").replace("#", "").split("\n")
        for r, line in enumerate(lines, start=1):
            cell = ws4.cell(row=r, column=1, value=line.strip())
            cell.alignment = Alignment(wrap_text=True)
            if line.startswith("##") or (line.strip() and r <= 3):
                cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"nubimetrics_{meta.get('category_name','categoria').replace(' ','_')}_{meta.get('period','')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/sales-summary")
@login_required
def sales_summary():
    from datetime import datetime, timezone, timedelta
    tz_arg = timezone(timedelta(hours=-3))
    now = datetime.now(tz_arg)

    user_id = get_my_user_id()
    if not user_id:
        return jsonify({"error": "No se pudo obtener el usuario de MeLi."}), 502

    fmt = '%Y-%m-%dT%H:%M:%S.000-0300'
    today_from  = now.replace(hour=0, minute=0, second=0, microsecond=0).strftime(fmt)
    month_from  = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime(fmt)
    now_str     = now.strftime(fmt)

    today_amt, today_cnt = fetch_orders_total(user_id, today_from, now_str)
    month_amt, month_cnt = fetch_orders_total(user_id, month_from, now_str)

    return jsonify({
        "today": {"amount": today_amt, "orders": today_cnt},
        "month": {"amount": month_amt, "orders": month_cnt},
        "as_of": now.strftime('%H:%M'),
    })


@app.route("/api/my-store")
@login_required
def my_store():
    items, error = get_my_store_items()
    if error:
        return jsonify({"error": error}), 502

    from collections import defaultdict
    items = [i for i in items if i["available_quantity"] > 0]
    active = [i for i in items if i["status"] == "active"]
    total_revenue = sum(i["revenue"] for i in items)
    cat_rev = defaultdict(int)
    for item in items:
        cat_rev[item["category_id"]] += item["revenue"]
    top_cats = sorted(cat_rev.items(), key=lambda x: -x[1])[:5]

    return jsonify({
        "items": items,
        "total": len(items),
        "active_count": len(active),
        "total_revenue_est": total_revenue,
        "top_categories": [{"cat": c, "revenue": r} for c, r in top_cats],
    })


@app.route("/api/my-store-analyze", methods=["POST"])
@login_required
def my_store_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    items = data.get("items", [])
    target_revenue = data.get("target_revenue", 30_000_000)
    if not items:
        return jsonify({"error": "Sin datos de tienda para analizar."}), 400

    top_items = items[:50]
    total_rev = sum(i["revenue"] for i in items)
    active_count = sum(1 for i in items if i["status"] == "active")

    summary = {
        "total_publicaciones": len(items),
        "publicaciones_activas": active_count,
        "revenue_historico_total_ARS": total_rev,
        "objetivo_incremento_mensual_ARS": target_revenue,
        "top_50_por_revenue_historico": [
            {
                "titulo": i["title"],
                "precio_ARS": i["price"],
                "unidades_vendidas_historico": i["sold_quantity"],
                "revenue_estimado_ARS": i["revenue"],
                "estado": i["status"],
                "categoria_id": i["category_id"],
            }
            for i in top_items
        ],
    }

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso el portfolio completo de un vendedor activo en MeLi Argentina.\n\n"
        f"El vendedor tiene {len(items)} publicaciones ({active_count} activas) "
        f"con un revenue histórico estimado de ${total_rev:,.0f} ARS.\n"
        f"Su objetivo concreto es sumar ${target_revenue:,.0f} ARS MÁS por mes.\n\n"
        "Nota: 'revenue_estimado' = precio × unidades_vendidas_historico (total histórico, no mensual).\n"
        "Usá los precios y productos como referencia de categoría y ticket promedio.\n\n"
        "Analizá su portfolio y recomendá 1 a 3 productos NUEVOS que debería agregar "
        "para alcanzar ese objetivo de crecimiento mensual.\n\n"
        "Para cada producto recomendado:\n"
        "- **Producto**: nombre/descripción específica y variante\n"
        "- **Por qué encaja**: relación con lo que ya vende y demanda del mercado\n"
        "- **Precio de venta sugerido** (ARS)\n"
        "- **Unidades mensuales estimadas** para alcanzar el objetivo\n"
        "- **Revenue mensual potencial** de ese producto\n"
        "- **Veredicto**: 🟢 Alta probabilidad / 🟡 Evaluar / 🔴 Riesgo alto\n\n"
        "Al final, un **Resumen ejecutivo**:\n"
        f"- Revenue mensual potencial total de los productos recomendados vs objetivo ${target_revenue:,.0f} ARS\n"
        "- Qué producto arrancar primero y por qué\n\n"
        f"Portfolio:\n{json.dumps(summary, ensure_ascii=False, indent=1)}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/sourcing-report")
@login_required
def sourcing_report():
    return render_template("sourcing_report.html")


@app.route("/api/sourcing-analyze", methods=["POST"])
@login_required
def sourcing_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    products = data.get("products", [])
    target_revenue = data.get("target_revenue", 0)
    min_products = data.get("min_products", 1)
    max_products = data.get("max_products", 3)
    shipping = data.get("shipping", "courier")
    tc = data.get("tc", 1500)

    if not products:
        return jsonify({"error": "Sin datos de productos para analizar."}), 400

    shipping_label = "Courier (aéreo)" if shipping == "courier" else "Marítimo (contenedor)"
    shipping_criteria = (
        "COURIER (aéreo): priorizar productos livianos (<1kg), compactos, alto valor/peso. "
        "Evitar productos voluminosos o pesados. Ideal para electrónica pequeña, accesorios, sensores, módulos."
        if shipping == "courier"
        else
        "MARÍTIMO: pueden ser productos más grandes o pesados (hasta 20-30kg, voluminosos). "
        "Mayor plazo de entrega (30-45 días). Ideal para herramientas, equipos, muebles, productos de mayor volumen."
    )

    products_str = json.dumps(products[:50], ensure_ascii=False, indent=1)

    prompt = (
        f"Sos un experto en sourcing y e-commerce en MercadoLibre Argentina.\n"
        f"Te paso datos reales de demanda de Nubimetrics (ventas históricas de MeLi Argentina).\n\n"
        f"CRITERIOS DEL VENDEDOR:\n"
        f"- Objetivo de facturación mensual adicional: ${target_revenue:,.0f} ARS\n"
        f"- Cantidad de productos a lanzar: entre {min_products} y {max_products}\n"
        f"- Método de importación: {shipping_label}\n"
        f"  → {shipping_criteria}\n"
        f"- Tipo de cambio referencia: ${tc:,.0f} ARS/USD\n\n"
        f"DATOS DE MERCADO (productos agrupados por título, ordenados por demanda total):\n"
        f"Cada producto incluye: título, precio_promedio_ARS, total_unidades, total_revenue_ARS, "
        f"vendedores_únicos, pct_full, categoría, archivo_fuente.\n\n"
        f"{products_str}\n\n"
        f"TAREA:\n"
        f"Seleccioná los mejores {min_products} a {max_products} productos de esta lista para "
        f"que el vendedor los importe y venda en MeLi, cumpliendo:\n"
        f"1. Alcanzar ${target_revenue:,.0f} ARS/mes adicionales en total\n"
        f"2. Apto para importar vía {shipping_label}\n"
        f"3. Competencia manejable (no dominada por pocos vendedores con alta concentración)\n"
        f"4. Demanda probada en datos reales de Nubimetrics\n\n"
        f"Para cada producto recomendado usá este formato:\n"
        f"### [Nombre del producto]\n"
        f"- **Demanda del mercado**: unidades totales vendidas, revenue total, vendedores compitiendo\n"
        f"- **Mi captura estimada**: si capturo X% del mercado = Y unidades/mes = $Z ARS/mes\n"
        f"- **Precio de venta sugerido**: $X ARS\n"
        f"- **FOB estimado**: USD X–Y por unidad (China)\n"
        f"- **Apto para {shipping_label}**: peso/tamaño estimado y por qué aplica\n"
        f"- **Veredicto**: 🟢 Alta oportunidad / 🟡 Evaluar / 🔴 Evitar\n\n"
        f"Al final, una tabla resumen:\n"
        f"| Producto | Precio sugerido | Captura estimada/mes |\n"
        f"| --- | --- | --- |\n"
        f"Con una línea final indicando si el objetivo de ${target_revenue:,.0f} ARS es alcanzable."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


# ─── Competidores ─────────────────────────────────────────────────────────

def _parse_seller_input(raw):
    """Accepts numeric ID, _CustId_ URL, perfil URL or plain alias.
    Returns (seller_id, error_msg)."""
    raw = (raw or "").strip()
    if not raw:
        return None, "Ingresá un alias, ID o URL del vendedor."
    if raw.isdigit():
        return raw, None
    m = re.search(r"_CustId_(\d+)", raw)
    if m:
        return m.group(1), None
    m = re.search(r"/perfil/([\w.\-]+)", raw)
    alias = m.group(1) if m else raw
    if not re.fullmatch(r"[\w.\-]+", alias):
        return None, "No pude interpretar ese vendedor. Pegá la URL de su perfil o su ID numérico."
    seller_id = resolve_seller_alias(alias)
    if seller_id:
        return seller_id, None
    return None, (
        "No pude resolver el alias automáticamente (MeLi bloquea la consulta desde el servidor). "
        "Abrí el perfil del vendedor en MeLi, tocá 'Ver más datos' o su listado de productos, "
        "y pegá acá esa URL (contiene _CustId_) — o usá el botón ➕ desde los resultados de búsqueda."
    )


@app.route("/api/competitors")
@login_required
def competitors_list():
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT c.*,
                   (SELECT COUNT(*) FROM competitor_snapshots s WHERE s.competitor_id = c.id) AS snapshots,
                   (SELECT MAX(taken_at) FROM competitor_snapshots s WHERE s.competitor_id = c.id) AS last_snapshot,
                   (SELECT item_count FROM competitor_snapshots s WHERE s.competitor_id = c.id
                    ORDER BY taken_at DESC LIMIT 1) AS item_count
            FROM competitors c ORDER BY c.added_at DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/competitors", methods=["POST"])
@login_required
def competitors_add():
    data = request.get_json() or {}
    seller_id, err = _parse_seller_input(data.get("seller"))
    if err:
        return jsonify({"error": err}), 422

    info = get_seller_info(seller_id)
    if not info:
        return jsonify({"error": f"MeLi no devolvió datos para el vendedor {seller_id}."}), 404

    with sqlite3.connect(DB_PATH) as conn:
        try:
            cur = conn.execute(
                "INSERT INTO competitors (seller_id, nickname, permalink) VALUES (?, ?, ?)",
                (info["seller_id"], info["nickname"], info["permalink"]),
            )
        except sqlite3.IntegrityError:
            return jsonify({"error": f"Ya estás siguiendo a {info['nickname']}."}), 409
    return jsonify({"id": cur.lastrowid, **info})


@app.route("/api/competitors/<int:cid>", methods=["DELETE"])
@login_required
def competitors_delete(cid):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM competitor_snapshots WHERE competitor_id = ?", (cid,))
        conn.execute("DELETE FROM competitors WHERE id = ?", (cid,))
    return jsonify({"ok": True})


def _set_job(cid, status, pct=0, msg="", error=""):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "INSERT INTO competitor_jobs (competitor_id, status, pct, msg, error, updated_at)"
            " VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)"
            " ON CONFLICT(competitor_id) DO UPDATE SET status=excluded.status, pct=excluded.pct,"
            " msg=excluded.msg, error=excluded.error, updated_at=CURRENT_TIMESTAMP",
            (cid, status, pct, msg, error),
        )


def _run_snapshot_job(cid, seller_id, max_items):
    try:
        result = snapshot_seller(
            seller_id, max_items,
            progress=lambda pct, msg: _set_job(cid, "running", pct, msg),
        )
        if "error" in result:
            _set_job(cid, "error", error=result["error"])
            return
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT INTO competitor_snapshots (competitor_id, item_count, seller_json, items_json)"
                " VALUES (?, ?, ?, ?)",
                (cid, len(result["items"]),
                 json.dumps(result["seller"], ensure_ascii=False),
                 json.dumps(result["items"], ensure_ascii=False)),
            )
        _set_job(cid, "done", 100, f"{len(result['items'])} publicaciones")
    except Exception as e:
        logging.exception("snapshot job failed for competitor %s", cid)
        _set_job(cid, "error", error=str(e))


# Sin señales de progreso durante este tiempo, el job se considera muerto
# (deploy/restart de Railway mata el thread sin actualizar la tabla).
_JOB_STALE_SECS = 600


def _get_job(cid):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        job = conn.execute(
            "SELECT *, (julianday(CURRENT_TIMESTAMP) - julianday(updated_at)) * 86400 AS age"
            " FROM competitor_jobs WHERE competitor_id = ?", (cid,),
        ).fetchone()
    if not job:
        return {"status": "idle"}
    d = dict(job)
    if d["status"] == "running" and d["age"] > _JOB_STALE_SECS:
        d["status"] = "error"
        d["error"] = "El snapshot se interrumpió (probable reinicio del servidor). Corré uno nuevo."
    d.pop("age", None)
    return d


@app.route("/api/competitors/<int:cid>/snapshot", methods=["POST"])
@login_required
def competitors_snapshot(cid):
    import threading

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        comp = conn.execute("SELECT * FROM competitors WHERE id = ?", (cid,)).fetchone()
    if not comp:
        return jsonify({"error": "Competidor no encontrado."}), 404

    if _get_job(cid).get("status") == "running":
        return jsonify({"error": "Ya hay un snapshot en curso para este competidor."}), 409

    max_items = min(int((request.get_json(silent=True) or {}).get("max_items", 0)), 5000)
    _set_job(cid, "running", 1, "Iniciando…")
    threading.Thread(
        target=_run_snapshot_job, args=(cid, comp["seller_id"], max_items), daemon=True,
    ).start()
    return jsonify({"started": True}), 202


@app.route("/api/competitors/<int:cid>/snapshot-status")
@login_required
def competitors_snapshot_status(cid):
    return jsonify(_get_job(cid))


def _item_key(item):
    return item.get("item_id") or item.get("catalog_id") or item.get("title")


def _build_competitor_report(comp, latest, previous):
    cur_items = json.loads(latest["items_json"])
    seller = json.loads(latest["seller_json"] or "null")
    prev_items = json.loads(previous["items_json"]) if previous else []
    prev_map = {_item_key(i): i for i in prev_items}
    cur_keys = {_item_key(i) for i in cur_items}

    price_changes = []
    for item in cur_items:
        old = prev_map.get(_item_key(item))
        if old and old.get("price") and item.get("price") and old["price"] != item["price"]:
            item["price_prev"] = old["price"]
            price_changes.append({
                "title": item["title"], "old": old["price"], "new": item["price"],
                "pct": round((item["price"] / old["price"] - 1) * 100, 1),
            })

    new_items = [i for i in cur_items if previous and _item_key(i) not in prev_map]
    removed_items = [
        {"title": i["title"], "price": i.get("price"), "v30": i.get("v30")}
        for i in prev_items if _item_key(i) not in cur_keys
    ]

    prev_seller = json.loads(previous["seller_json"] or "null") if previous else None
    transactions_delta = None
    if seller and prev_seller and seller.get("transactions_total") and prev_seller.get("transactions_total"):
        transactions_delta = seller["transactions_total"] - prev_seller["transactions_total"]

    return {
        "competitor": {"id": comp["id"], "nickname": comp["nickname"], "seller_id": comp["seller_id"],
                       "permalink": comp["permalink"]},
        "seller": seller,
        "taken_at": latest["taken_at"],
        "prev_taken_at": previous["taken_at"] if previous else None,
        "transactions_delta": transactions_delta,
        "items": sorted(cur_items, key=lambda i: -(i.get("v30") or 0)),
        "price_changes": sorted(price_changes, key=lambda c: abs(c["pct"]), reverse=True),
        "new_items": [{"title": i["title"], "price": i.get("price")} for i in new_items],
        "removed_items": removed_items,
        "totals": {
            "items": len(cur_items),
            "v30": sum(i.get("v30") or 0 for i in cur_items),
            "v30_prev": sum(i.get("v30_prev") or 0 for i in cur_items),
            "free_shipping_pct": round(100 * sum(1 for i in cur_items if i.get("free_shipping")) / len(cur_items)) if cur_items else 0,
        },
    }


@app.route("/api/competitors/<int:cid>/report")
@login_required
def competitors_report(cid):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        comp = conn.execute("SELECT * FROM competitors WHERE id = ?", (cid,)).fetchone()
        if not comp:
            return jsonify({"error": "Competidor no encontrado."}), 404
        snaps = conn.execute(
            "SELECT * FROM competitor_snapshots WHERE competitor_id = ? ORDER BY taken_at DESC LIMIT 2",
            (cid,),
        ).fetchall()
    if not snaps:
        return jsonify({"error": "Este competidor todavía no tiene snapshots. Corré uno primero."}), 404
    return jsonify(_build_competitor_report(comp, snaps[0], snaps[1] if len(snaps) > 1 else None))


@app.route("/competitor-report")
@login_required
def competitor_report():
    return render_template("competitor_report.html")


@app.route("/api/competitor-analyze", methods=["POST"])
@login_required
def competitor_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    report = request.get_json() or {}
    if not report.get("items"):
        return jsonify({"error": "Sin datos para analizar."}), 400

    # Compact payload: top 40 items by 30-day visits, weekly series trimmed to last 8 weeks
    slim_items = []
    for i in report["items"][:40]:
        slim_items.append({
            "titulo": i.get("title", "")[:80],
            "precio": i.get("price"),
            "precio_anterior": i.get("price_prev"),
            "envio_gratis": i.get("free_shipping"),
            "visitas_30d": i.get("v30"),
            "visitas_30d_previas": i.get("v30_prev"),
            "visitas_semanales": [w["v"] for w in (i.get("weekly") or [])[-8:]],
        })

    payload = {
        "vendedor": report.get("seller"),
        "snapshot_actual": report.get("taken_at"),
        "snapshot_anterior": report.get("prev_taken_at"),
        "delta_transacciones_historicas": report.get("transactions_delta"),
        "totales": report.get("totals"),
        "cambios_de_precio": report.get("price_changes", [])[:25],
        "publicaciones_nuevas": report.get("new_items", [])[:25],
        "publicaciones_dadas_de_baja": report.get("removed_items", [])[:25],
        "top_publicaciones": slim_items,
    }

    prompt = (
        "Sos un analista experto en e-commerce de MercadoLibre Argentina. "
        "Te paso datos de monitoreo de un vendedor COMPETIDOR: su catálogo actual con visitas "
        "diarias reales de MeLi (proxy de demanda; no hay datos de ventas), y si hay snapshot "
        "anterior, los cambios de precio, altas y bajas de publicaciones.\n\n"
        f"DATOS:\n{json.dumps(payload, ensure_ascii=False, indent=1)}\n\n"
        "Notas sobre los datos:\n"
        "- visitas_semanales: serie de visitas por semana (más antigua → más reciente), sirve para detectar tendencia.\n"
        "- delta_transacciones_historicas: ventas reales aproximadas entre snapshots (si existe).\n"
        "- Si no hay snapshot anterior, enfocate en la foto actual y las tendencias de visitas.\n\n"
        "TAREA — respondé en markdown con estas secciones:\n"
        "## Perfil del competidor\n"
        "Qué vende, en qué segmento de precios juega, su escala y reputación. 3-4 líneas.\n"
        "## Productos calientes\n"
        "### por cada producto destacado (máx 5): qué está traccionando (visitas y tendencia), "
        "precio, y veredicto 🟢 atacar / 🟡 observar / 🔴 no conviene competir.\n"
        "## Movimientos recientes\n"
        "Cambios de precio relevantes, publicaciones nuevas y dadas de baja, y qué revelan de su "
        "estrategia. Si no hay snapshot anterior, indicá que esta sección estará disponible a partir "
        "del próximo snapshot.\n"
        "## Lectura estratégica\n"
        "Hipótesis sobre su estrategia (líder de precio, surtido amplio, nichos premium, etc.).\n"
        "## Acciones recomendadas\n"
        "3 a 5 acciones concretas y priorizadas para competir contra este vendedor.\n\n"
        "Usá formatos abreviados para números grandes ($1.2M, 14.4k). Sé directo y accionable."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
